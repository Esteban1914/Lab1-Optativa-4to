from openpyxl import load_workbook

class Suppliers:   
    def __init__(self,xlsx_dir):
        suppliers={}
        wb=load_workbook(xlsx_dir)        
        ws = wb.active
        suppliers_col = ws['B']
        for row,supplier_cell in enumerate(suppliers_col,start=1):
            if row == 1:
                continue
            supplier_name=supplier_cell.value.__str__()
            if supplier_name not in suppliers: 
                suppliers.update({supplier_name : Supplier(supplier_name)})
            product=Product(
                n_product        = ws[ "A{data}".format(data=row)].value,
                purchased_amount = ws[ "C{data}".format(data=row)].value,
                unit_cost        = ws[ "D{data}".format(data=row)].value,
                sale_price       = ws[ "E{data}".format(data=row)].value,
                quantity_sold    = ws[ "F{data}".format(data=row)].value
            )
            suppliers[supplier_name].addProduct(product)    
        self.__suppliers=suppliers     
           
    def show_unsold(self):
        print("Invientario no vendido")
        total_value=0
        total_products=0
        for supplier_value in self.__suppliers.values():
            t_p, t_v = supplier_value.show_unsold()
            total_products += t_p
            total_value += t_v
        print("Total de productos sin vender:{}\nValor Total:{}".format(total_products,total_value))
 
# Another Internal Class

class InventaryError(Exception):
        pass
    
class Product:
    def __init__(self,n_product,purchased_amount,quantity_sold,unit_cost,sale_price) -> None:
        self.n_product=n_product
        self.purchased_amount=purchased_amount
        self.quantity_sold=quantity_sold
        self.unit_cost=unit_cost
        self.sale_price=sale_price
        
    def show_unsold(self):
        try:
            in_inventary=self.purchased_amount - self.quantity_sold
            if in_inventary < 0:
                raise InventaryError("\t\tError, la cantidad vendida({}) es mayor q la comprada({})".format(self.quantity_sold,self.purchased_amount))
            if in_inventary:
                value=in_inventary * self.unit_cost
                print("\t\tProducto:",self.n_product)
                print("\t\t\tNo vendidos:{}\n\t\t\tValor:{}".format(
                    in_inventary,
                    value 
                    ))
                return in_inventary,value
            else:
                print("\t\tTodos los productos vendidos")
        except InventaryError as e:
            print(e.args[0])
        return 0,0

class Supplier:
    def __init__(self,name):
        self.name=name
        self.products=[]
        
    def addProduct(self,product):
        self.products.append(product)
        
    def show_unsold(self):
        total_value=0
        total_products=0
        print("\tSuministrador: ",self.name)
        for product in self.products:
            t_p, t_v = product.show_unsold()
            total_products += t_p
            total_value += t_v
        return total_products,total_value